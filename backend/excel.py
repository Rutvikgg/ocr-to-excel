"""
This is where we implement excel functionality
"""

from openpyxl import Workbook, load_workbook

import backend.constants as c
import backend.functions as f


class ExcelWriter:
    def __init__(self) -> None:
        pass

    @staticmethod
    def create_wb(data: dict, save_location: str | None = None) -> None:
        # Creating workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = c.SINGLE_SHEET_TITLE

        # Adding heading for columns
        f.add_headers([sheet], [c.SINGLE_SHEET_HEADERS])

        # Adding Data
        line_items = f.get_line_items(data.get("line_items"))
        append_list = [data.get("date"),
                       data.get("time"),
                       data.get("locale", {}).get("language"),
                       data.get("locale", {}).get("currency"),
                       data.get("document_type"),
                       data.get("invoice_number"),
                       f.get_ref_nums(data.get("reference_numbers")),
                       data.get("due_date"),
                       data.get("total_net"),
                       data.get("total_tax"),
                       data.get("total_amount"),
                       data.get("tip"),
                       line_items[0][0],
                       line_items[0][1],
                       line_items[0][2],
                       line_items[0][3],
                       line_items[0][4],
                       line_items[0][5],
                       line_items[0][6],
                       data.get("supplier", {}).get("name"),
                       data.get("supplier", {}).get("phone_number"),
                       data.get("supplier", {}).get("address"),
                       f.get_reg_info(data.get("supplier", {}).get("reg_info")),
                       f.get_payment_details(data.get("supplier", {}).get("payment_details")),
                       data.get("customer", {}).get("name"),
                       data.get("customer", {}).get("address"),
                       f.get_reg_info(data.get("customer", {}).get("reg_info"))]
        sheet.append(append_list)

        for idx, item in enumerate(line_items[1:]):
            sheet.cell(row=idx + 5, column=13).value = item[0]
            sheet.cell(row=idx + 5, column=14).value = item[1]
            sheet.cell(row=idx + 5, column=15).value = item[2]
            sheet.cell(row=idx + 5, column=16).value = item[3]
            sheet.cell(row=idx + 5, column=17).value = item[4]
            sheet.cell(row=idx + 5, column=18).value = item[5]
            sheet.cell(row=idx + 5, column=19).value = item[6]

        # Cell formatting
        f.format_cell_width(sheet)

        # Saving worksheet
        workbook.save(save_location)

    @staticmethod
    def create_multi_sheet_wb(data: dict, save_location: str | None = None) -> None:
        # Creating workbook and sheets
        workbook = Workbook()
        main_info_sheet = workbook.active
        main_info_sheet.title = c.MULTI_SHEET_TITLES[0]
        supplier_info_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[1])
        customer_info_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[2])
        line_items_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[3])

        # Adding heading for columns in sheets
        f.add_headers([main_info_sheet, supplier_info_sheet, customer_info_sheet, line_items_sheet],
                      [c.MAIN_SHEET_HEADERS, c.SUPPLIER_SHEET_HEADERS, c.CUSTOMER_SHEET_HEADERS,
                       c.LINE_ITEMS_SHEET_HEADERS])

        # Adding Data
        main_info_sheet.append([data.get("date"),
                                data.get("time"),
                                data.get("locale", {}).get("language"),
                                data.get("locale", {}).get("currency"),
                                data.get("document_type"),
                                data.get("invoice_number"),
                                f.get_ref_nums(data.get("reference_numbers")),
                                data.get("due_date"),
                                data.get("total_net"),
                                data.get("total_tax"),
                                data.get("total_amount"),
                                data.get("tip")])

        supplier_info_sheet.append([data.get("invoice_number"),
                                    f.get_payment_details(data.get("supplier", {}).get("payment_details")),
                                    data.get("supplier", {}).get("name"),
                                    data.get("supplier", {}).get("address"),
                                    data.get("supplier", {}).get("phone_number"),
                                    f.get_reg_info(data.get("supplier", {}).get("reg_info"))])

        customer_info_sheet.append([data.get("invoice_number"),
                                    data.get("customer", {}).get("name"),
                                    data.get("customer", {}).get("address"),
                                    f.get_reg_info(data.get("customer", {}).get("reg_info"))])

        line_items = f.get_line_items(data.get("line_items"))
        line_items_sheet.append([data.get("invoice_number"),
                                 line_items[0][0],
                                 line_items[0][1],
                                 line_items[0][2],
                                 line_items[0][3],
                                 line_items[0][4],
                                 line_items[0][5],
                                 line_items[0][6]])
        for idx, item in enumerate(line_items[1:]):
            line_items_sheet.cell(row=idx + 5, column=2).value = item[0]
            line_items_sheet.cell(row=idx + 5, column=3).value = item[1]
            line_items_sheet.cell(row=idx + 5, column=4).value = item[2]
            line_items_sheet.cell(row=idx + 5, column=5).value = item[3]
            line_items_sheet.cell(row=idx + 5, column=6).value = item[4]
            line_items_sheet.cell(row=idx + 5, column=7).value = item[5]
            line_items_sheet.cell(row=idx + 5, column=8).value = item[6]

        # Cell formatting
        f.format_cell_width(main_info_sheet, supplier_info_sheet, customer_info_sheet, line_items_sheet)

        # Saving workbook
        workbook.save(save_location)

    @staticmethod
    def is_wb_single_sheet(wb_location: str) -> bool:
        workbook = load_workbook(wb_location)
        sheets = workbook.sheetnames
        if c.SINGLE_SHEET_TITLE in sheets:
            return True
        else:
            return False

    @staticmethod
    def append_wb(data: dict, wb_location: str, save_location: str | None = None) -> None:
        workbook = load_workbook(wb_location)
        try:
            sheet = workbook[c.SINGLE_SHEET_TITLE]
        except KeyError:
            sheet = workbook.create_sheet(c.SINGLE_SHEET_TITLE)

        # Adding Data
        line_items = f.get_line_items(data.get("line_items"))
        append_list = [data.get("date"),
                       data.get("time"),
                       data.get("locale", {}).get("language"),
                       data.get("locale", {}).get("currency"),
                       data.get("document_type"),
                       data.get("invoice_number"),
                       f.get_ref_nums(data.get("reference_numbers")),
                       data.get("due_date"),
                       data.get("total_net"),
                       data.get("total_tax"),
                       data.get("total_amount"),
                       data.get("tip"),
                       line_items[0][0],
                       line_items[0][1],
                       line_items[0][2],
                       line_items[0][3],
                       line_items[0][4],
                       line_items[0][5],
                       line_items[0][6],
                       data.get("supplier", {}).get("name"),
                       data.get("supplier", {}).get("phone_number"),
                       data.get("supplier", {}).get("address"),
                       f.get_reg_info(data.get("supplier", {}).get("reg_info")),
                       f.get_payment_details(data.get("supplier", {}).get("payment_details")),
                       data.get("customer", {}).get("name"),
                       data.get("customer", {}).get("address"),
                       f.get_reg_info(data.get("customer", {}).get("reg_info"))]
        sheet.append(append_list)
        max_row = sheet.max_row
        for idx, item in enumerate(line_items[1:]):
            sheet.cell(row=idx + max_row + 1, column=13).value = item[0]
            sheet.cell(row=idx + max_row + 1, column=14).value = item[1]
            sheet.cell(row=idx + max_row + 1, column=15).value = item[2]
            sheet.cell(row=idx + max_row + 1, column=16).value = item[3]
            sheet.cell(row=idx + max_row + 1, column=17).value = item[4]
            sheet.cell(row=idx + max_row + 1, column=18).value = item[5]
            sheet.cell(row=idx + max_row + 1, column=19).value = item[6]

        # Cell formatting
        f.format_cell_width(sheet)

        # Saving workbook
        if save_location is None:
            workbook.save(wb_location)
        else:
            workbook.save(save_location)

    @staticmethod
    def append_multi_sheet_wb(data: dict, wb_location: str, save_location: str | None = None) -> None:
        workbook = load_workbook(wb_location)
        try:
            main_info_sheet = workbook[c.MULTI_SHEET_TITLES[0]]
            supplier_info_sheet = workbook[c.MULTI_SHEET_TITLES[1]]
            customer_info_sheet = workbook[c.MULTI_SHEET_TITLES[2]]
            line_items_sheet = workbook[c.MULTI_SHEET_TITLES[3]]
        except KeyError:
            main_info_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[0])
            supplier_info_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[1])
            customer_info_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[2])
            line_items_sheet = workbook.create_sheet(c.MULTI_SHEET_TITLES[3])

        # Adding Data
        main_info_sheet.append([data.get("date"),
                                data.get("time"),
                                data.get("locale", {}).get("language"),
                                data.get("locale", {}).get("currency"),
                                data.get("document_type"),
                                data.get("invoice_number"),
                                f.get_ref_nums(data.get("reference_numbers")),
                                data.get("due_date"),
                                data.get("total_net"),
                                data.get("total_tax"),
                                data.get("total_amount"),
                                data.get("tip")])

        supplier_info_sheet.append([data.get("invoice_number"),
                                    f.get_payment_details(data.get("supplier", {}).get("payment_details")),
                                    data.get("supplier", {}).get("name"),
                                    data.get("supplier", {}).get("address"),
                                    data.get("supplier", {}).get("phone_number"),
                                    f.get_reg_info(data.get("supplier", {}).get("reg_info"))])

        customer_info_sheet.append([data.get("invoice_number"),
                                    data.get("customer", {}).get("name"),
                                    data.get("customer", {}).get("address"),
                                    f.get_reg_info(data.get("customer", {}).get("reg_info"))])

        line_items = f.get_line_items(data.get("line_items"))
        line_items_sheet.append([data.get("invoice_number"),
                                 line_items[0][0],
                                 line_items[0][1],
                                 line_items[0][2],
                                 line_items[0][3],
                                 line_items[0][4],
                                 line_items[0][5],
                                 line_items[0][6]])
        max_row = line_items_sheet.max_row
        for idx, item in enumerate(line_items[1:]):
            line_items_sheet.cell(row=idx + max_row + 1, column=2).value = item[0]
            line_items_sheet.cell(row=idx + max_row + 1, column=3).value = item[1]
            line_items_sheet.cell(row=idx + max_row + 1, column=4).value = item[2]
            line_items_sheet.cell(row=idx + max_row + 1, column=5).value = item[3]
            line_items_sheet.cell(row=idx + max_row + 1, column=6).value = item[4]
            line_items_sheet.cell(row=idx + max_row + 1, column=7).value = item[5]
            line_items_sheet.cell(row=idx + max_row + 1, column=8).value = item[6]

        # Cell formatting
        f.format_cell_width(main_info_sheet, supplier_info_sheet, customer_info_sheet, line_items_sheet)

        # Saving workbook
        if save_location is None:
            workbook.save(wb_location)
        else:
            workbook.save(save_location)
