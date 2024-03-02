import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

workbook = load_workbook("../test_data/excel_test.xlsx")

data = ["Name", "Address", "Contact"]
data2 = ["Name", None, "Contact"]
sheet = workbook["Line Items"]
row = 3
# workbook.create_sheet("Description")
# sheet.title = "Line Items"
# bold_font = Font(bold=True)
# for col_idx, header in enumerate(data, 1):
#     sheet.cell(row, col_idx).value = header
#     sheet.cell(row, col_idx).font = bold_font

# for i, d in enumerate(data):
#     print(i, d)

# sheet.append(data2)
# workbook.save("../test_data/excel_test.xlsx")

# print(sheet.cell(3, 1).value)

# for col in sheet.iter_cols(min_col=1, max_col=3):
#     print(col[0].col_idx)
#     for cell in col:
#         print(cell.value)

fin_doc_data = {
    "document_type": "Expense Receipt",
    "locale": {
        "language": "en",
        "currency": "INR"
    },
    "invoice_number": 80456,
    "reference_numbers": ["AP10256", "UN5847", "AT548R"],
    "date": "02-03-2024",
    "time": "08:05 PM",
    "due_date": "06-04-2024",
    "total_net": 42000,
    "total_tax": 456.23,
    "total_amount": 42456.23,
    "tip": None,
    "taxes": [{"value": 456, "rate": 3.5}, {"value": 456, "rate": 3.5}],
    "supplier": {
        "payment_details": [{
            "iban": " GB29 NWBK 6016 1331 9268 19",
            "swift": "NWBKGB2L",
            "account_number": 1234567890,
            "routing_number": 123456789
        }, {
            "iban": " GB29 NWBK 6016 1331 9268 19",
            "swift": "NWBKGB2L",
            "account_number": 3658974106,
            "routing_number": 12569843
        }],
        "name": "John Doe Enterprises",
        "reg_info": [{
            "type": "Fake Business Registry",
            "value": "ABC123456"
        }],
        "address": "123 Fake Street, Faketown, FK1234",
        "phone_number": "+1 (555) 123-4567"
    },
    "customer": {
        "name": "Jane Smith Co.",
        "reg_info": [{
            "type": "Fake Business Registry",
            "value": "ABC123456"
        }],
        "address": "456 Imaginary Avenue, Fantasyville, FA5678"
    },
    "line_items": [
        {
            "product_code": "PRD001",
            "description": "Widget A",
            "quantity": 10,
            "unit_price": 25.99,
            "total_amount": 259.90,
            "tax_rate": 0.1,
            "tax_amount": 25.99
        },
        {
            "product_code": "PRD002",
            "description": "Widget B",
            "quantity": 5,
            "unit_price": 15.49,
            "total_amount": 77.45,
            "tax_rate": 0.08,
            "tax_amount": 6.20
        },
        {
            "product_code": "PRD003",
            "description": "Widget C",
            "quantity": 20,
            "unit_price": 8.75,
            "total_amount": 175.00,
            "tax_rate": 0.05,
            "tax_amount": 8.75
        }
    ]
}
items = list("RUTVIK")
for idx, item in enumerate(items[1:]):
    print(idx+2, item)
